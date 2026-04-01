#!/usr/bin/env python3
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill

FILE1 = "/Users/hyijie/Desktop/Interest_try_test/google_scholar_mcp/我在这里我在这里_glm5.xlsx"
FILE2 = "/Users/hyijie/Desktop/Interest_try_test/google_scholar_mcp/我在这里我在这里_part2.xlsx"
OUTPUT = "/Users/hyijie/Desktop/Interest_try_test/google_scholar_mcp/最后我在这里.xlsx"

def merge_excel_files():
    wb = Workbook()
    ws = wb.active
    ws.title = "Citations"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    headers = ["Title", "Year", "Publisher", "Abstract", "URL"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    current_row = 2
    total_papers = 0
    seen_titles = set()
    
    for filepath in [FILE1, FILE2]:
        try:
            wb_src = load_workbook(filepath)
            ws_src = wb_src.active
            
            for row in range(2, ws_src.max_row + 1):
                title = ws_src.cell(row=row, column=1).value
                if not title or title in seen_titles:
                    continue
                seen_titles.add(title)
                
                for col in range(1, 6):
                    value = ws_src.cell(row=row, column=col).value
                    ws.cell(row=current_row, column=col, value=value)
                
                ws.row_dimensions[current_row].height = 60
                current_row += 1
                total_papers += 1
            
            wb_src.close()
            print(f"Loaded: {filepath}")
        except Exception as e:
            print(f"Error loading {filepath}: {e}")
    
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 80
    ws.column_dimensions["E"].width = 40
    
    wb.save(OUTPUT)
    print(f"\nTotal unique papers: {total_papers}")
    print(f"Saved to: {OUTPUT}")

if __name__ == "__main__":
    merge_excel_files()
