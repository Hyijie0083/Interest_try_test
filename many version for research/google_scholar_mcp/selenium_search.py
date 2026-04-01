#!/usr/bin/env python3
import time
import random
import re
import sys
import os

import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

PAPER_TITLE = "A supramodal accumulation-to-bound signal that determines perceptual decisions in humans"
OUTPUT_FILENAME = "我在这里我在这里_glm5.xlsx"
MAX_RESULTS = 100


def random_delay(min_sec: float = 2.0, max_sec: float = 5.0):
    delay = random.uniform(min_sec, max_sec)
    time.sleep(delay)


def human_like_delay():
    delays = [
        (1.0, 2.0),
        (2.0, 4.0),
        (3.0, 6.0),
        (1.5, 3.5),
    ]
    min_d, max_d = random.choice(delays)
    time.sleep(random.uniform(min_d, max_d))


def setup_driver():
    options = uc.ChromeOptions()
    
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--lang=en-US")
    
    prefs = {
        "profile.default_content_setting_values.notifications": 2,
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False,
    }
    options.add_experimental_option("prefs", prefs)
    
    driver = uc.Chrome(options=options, version_main=146)
    
    return driver


def extract_year(text: str) -> str | None:
    year_match = re.search(r"\b(19|20)\d{2}\b", text)
    if year_match:
        return year_match.group()
    return None


def parse_citation_count(text: str) -> int:
    match = re.search(r"[\d,]+", text.replace(",", ""))
    if match:
        return int(match.group().replace(",", ""))
    return 0


def is_captcha_page(driver) -> bool:
    current_url = driver.current_url.lower()
    page_source = driver.page_source.lower()
    
    if "sorry" in current_url:
        return True
    if "sorry" in page_source and ("captcha" in page_source or "unusual traffic" in page_source):
        return True
    if "recaptcha" in page_source:
        return True
    if "verify" in current_url:
        return True
    
    return False


def wait_for_captcha_solve(driver, timeout: int = 300):
    print("\n" + "=" * 60)
    print("CAPTCHA DETECTED!")
    print("Please solve the captcha in the browser window.")
    print("The script will continue automatically after you solve it.")
    print("=" * 60 + "\n")
    
    start_time = time.time()
    while time.time() - start_time < timeout:
        if not is_captcha_page(driver):
            print("Captcha solved! Continuing...")
            return True
        time.sleep(2)
    
    return False


def search_paper(driver, paper_title: str) -> tuple[int, str | None]:
    search_url = f"https://scholar.google.com/scholar?q={paper_title.replace(' ', '+')}&hl=en&as_sdt=0,5"
    
    print(f"Navigating to Google Scholar...")
    driver.get(search_url)
    human_like_delay()
    
    if is_captcha_page(driver):
        if not wait_for_captcha_solve(driver):
            print("Captcha not solved within timeout")
            return 0, None
    
    cited_by = 0
    citation_id = None
    
    try:
        cited_by_links = driver.find_elements(By.PARTIAL_LINK_TEXT, "Cited by")
        if cited_by_links:
            link = cited_by_links[0]
            href = link.get_attribute("href")
            if href and "cites=" in href:
                id_match = re.search(r"cites=([^&]+)", href)
                if id_match:
                    citation_id = id_match.group(1)
            
            cited_text = link.text
            cited_by = parse_citation_count(cited_text)
            print(f"Found: Cited by {cited_by} papers")
    except NoSuchElementException:
        print("No 'Cited by' link found")
    
    return cited_by, citation_id


def get_citing_papers(driver, citation_id: str, max_results: int = 100) -> list[dict]:
    all_papers = []
    start = 0
    page_num = 0
    
    while len(all_papers) < max_results:
        cites_url = f"https://scholar.google.com/scholar?cites={citation_id}&as_sdt=2005&sciodt=0,5&hl=en&start={start}"
        
        print(f"\nFetching page {page_num + 1}...")
        driver.get(cites_url)
        human_like_delay()
        
        if is_captcha_page(driver):
            if not wait_for_captcha_solve(driver):
                print("Could not solve captcha, stopping...")
                break
        
        results = driver.find_elements(By.CSS_SELECTOR, "div.gs_r")
        
        if not results:
            print("No more results found")
            break
        
        for result in results:
            try:
                title_elem = result.find_element(By.CSS_SELECTOR, "h3.gs_rt a")
                title = title_elem.text.strip()
                url = title_elem.get_attribute("href")
                
                try:
                    abstract_elem = result.find_element(By.CSS_SELECTOR, "div.gs_rs")
                    abstract = abstract_elem.text.strip()
                except NoSuchElementException:
                    abstract = ""
                
                try:
                    info_elem = result.find_element(By.CSS_SELECTOR, "div.gs_a")
                    meta_text = info_elem.text.strip()
                except NoSuchElementException:
                    meta_text = ""
                
                year = extract_year(meta_text)
                
                publisher = None
                if meta_text:
                    parts = meta_text.split(" - ")
                    if parts:
                        publisher = parts[0].strip()
                
                all_papers.append({
                    "title": title,
                    "year": year,
                    "publisher": publisher,
                    "abstract": abstract,
                    "url": url,
                })
                
            except NoSuchElementException:
                continue
        
        print(f"Found {len(results)} papers on page {page_num + 1}, total: {len(all_papers)}")
        
        if len(results) < 10:
            break
        
        page_num += 1
        start += 10
        
        if page_num % 2 == 0:
            batch_delay = random.uniform(8, 15)
            print(f"Batch rest: {batch_delay:.1f}s...")
            time.sleep(batch_delay)
        else:
            random_delay(3, 6)
    
    return all_papers[:max_results]


def save_to_excel(papers: list[dict], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Citations"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    headers = ["Title", "Year", "Publisher", "Abstract", "URL"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    for row, paper in enumerate(papers, 2):
        ws.cell(row=row, column=1, value=paper["title"])
        ws.cell(row=row, column=2, value=paper["year"] or "N/A")
        ws.cell(row=row, column=3, value=paper["publisher"] or "N/A")
        ws.cell(row=row, column=4, value=paper["abstract"] if paper["abstract"] else "N/A")
        ws.cell(row=row, column=5, value=paper["url"] or "N/A")
        ws.row_dimensions[row].height = 60
    
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 80
    ws.column_dimensions["E"].width = 40
    
    wb.save(output_path)
    print(f"\nSaved {len(papers)} papers to {output_path}")


def main():
    print("=" * 60)
    print(f"Searching for: {PAPER_TITLE}")
    print(f"Output file: {OUTPUT_FILENAME}")
    print("=" * 60)
    
    driver = None
    try:
        print("\nInitializing Chrome browser (undetected mode)...")
        driver = setup_driver()
        
        print("\nStep 1: Searching for paper...")
        cited_by, citation_id = search_paper(driver, PAPER_TITLE)
        
        if not citation_id:
            print("Could not find citation ID. The paper may not have citations yet.")
            return
        
        print(f"\nCitation ID: {citation_id}")
        print(f"Cited by: {cited_by} papers")
        
        print("\nStep 2: Fetching citing papers...")
        papers = get_citing_papers(driver, citation_id, MAX_RESULTS)
        
        print(f"\nTotal papers collected: {len(papers)}")
        
        output_path = f"/Users/hyijie/Desktop/Interest_try_test/google_scholar_mcp/{OUTPUT_FILENAME}"
        save_to_excel(papers, output_path)
        
        print("\nDone!")
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        if driver:
            print("\nClosing browser...")
            driver.quit()


if __name__ == "__main__":
    main()
