#!/usr/bin/env python3
"""
MCP Server for Google Scholar.

This server provides tools to search Google Scholar, retrieve citations,
and export citation data to Excel files with keyword marking in title or abstract.
"""

import asyncio
import json
import random
import re
import time
from datetime import datetime
from typing import List, Optional

import httpx
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, ConfigDict, Field, field_validator

from mcp.server.fastmcp import FastMCP

mcp = FastMCP("google_scholar_mcp")

USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.2 Safari/605.1.15",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0",
]

HEADERS_TEMPLATE = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Sec-Fetch-User": "?1",
    "Cache-Control": "max-age=0",
    "TE": "trailers",
}


def _get_headers() -> dict:
    headers = HEADERS_TEMPLATE.copy()
    headers["User-Agent"] = random.choice(USER_AGENTS)
    return headers


def _jitter(base: float, multiplier: float = 1.0) -> float:
    return base + random.uniform(0, multiplier)


def _exponential_backoff_with_jitter(base: float, attempt: int, max_delay: float = 60.0) -> float:
    exp_delay = min(base * (2 ** attempt), max_delay)
    jitter = random.uniform(0, exp_delay * 0.5)
    return exp_delay + jitter


async def _smart_delay(last_request_time: float, min_interval: float = 3.0, max_interval: float = 8.0) -> float:
    elapsed = time.time() - last_request_time
    if elapsed < min_interval:
        delay = random.uniform(min_interval, max_interval)
        await asyncio.sleep(delay)
        return time.time()
    return last_request_time


class RateLimiter:
    def __init__(self):
        self.request_times: List[float] = []
        self.consecutive_errors: int = 0
        self.last_captcha_time: float = 0
        self.base_delay: float = 8.0
        self.max_delay: float = 180.0
        self.total_requests: int = 0
        self.successful_requests: int = 0
        self.session_cookies: dict = {}

    def record_success(self):
        self.consecutive_errors = 0
        self.successful_requests += 1
        self.request_times.append(time.time())
        if len(self.request_times) > 20:
            self.request_times.pop(0)

    def record_error(self):
        self.consecutive_errors += 1

    def record_captcha(self):
        self.consecutive_errors = 20
        self.last_captcha_time = time.time()

    def record_rate_limit(self):
        self.consecutive_errors = max(self.consecutive_errors + 5, 15)

    def get_recommended_delay(self) -> float:
        recent_requests = [t for t in self.request_times if time.time() - t < 120]
        if len(recent_requests) > 5:
            avg_gap = 120 / len(recent_requests) if recent_requests else 8
            base = max(avg_gap * 2.5, self.base_delay)
        else:
            base = self.base_delay

        if self.consecutive_errors == 0:
            return random.uniform(base, base * 1.5)

        delay = _exponential_backoff_with_jitter(
            base,
            min(self.consecutive_errors, 8),
            self.max_delay
        )
        return min(delay, self.max_delay)

    def get_adaptive_delay_for_page(self, page_num: int) -> float:
        if page_num < 3:
            return random.uniform(10.0, 20.0)
        elif page_num < 10:
            return random.uniform(15.0, 30.0)
        else:
            return random.uniform(20.0, 45.0)

    def should_wait_for_captcha_cooldown(self) -> bool:
        if self.last_captcha_time == 0:
            return False
        cooldown = time.time() - self.last_captcha_time
        return cooldown < 900

    def get_captcha_cooldown_remaining(self) -> float:
        if self.last_captcha_time == 0:
            return 0
        elapsed = time.time() - self.last_captcha_time
        return max(0, 900 - elapsed)


rate_limiter = RateLimiter()


def _is_captcha_page(response: httpx.Response) -> bool:
    if "sorry" in str(response.url).lower():
        return True
    if "sorry" in response.text.lower() and ("captcha" in response.text.lower() or "unusual traffic" in response.text.lower()):
        return True
    if response.status_code == 429:
        return True
    return False


def _is_rate_limited(response: httpx.Response) -> bool:
    if response.status_code == 429:
        return True
    if "too many requests" in response.text.lower() or "try again later" in response.text.lower():
        return True
    return False


class PaperInfo(BaseModel):
    title: str
    year: Optional[str] = None
    publisher: Optional[str] = None
    url: Optional[str] = None
    abstract: Optional[str] = None
    contains_keyword: bool = False


class SearchInput(BaseModel):
    model_config = ConfigDict(
        str_strip_whitespace=True,
        validate_assignment=True,
    )

    paper_title: str = Field(
        ...,
        description="Title of the paper to search on Google Scholar",
        min_length=1,
        max_length=500,
    )
    keyword: Optional[str] = Field(
        default=None,
        description="Keyword to mark papers (checks title AND abstract for this keyword)",
        max_length=100,
    )
    max_results: Optional[int] = Field(
        default=50,
        description="Maximum number of citation results to retrieve",
        ge=1,
        le=100,
    )
    output_filename: Optional[str] = Field(
        default=None,
        description="Custom output filename for the Excel file (without extension)",
        max_length=200,
    )

    @field_validator("paper_title")
    @classmethod
    def validate_paper_title(cls, v: str) -> str:
        if not v.strip():
            raise ValueError("Paper title cannot be empty")
        return v.strip()


def _parse_citation_count(text: str) -> int:
    match = re.search(r"[\d,]+", text.replace(",", ""))
    if match:
        return int(match.group().replace(",", ""))
    return 0


def _extract_year(text: str) -> Optional[str]:
    year_match = re.search(r"\b(19|20)\d{2}\b", text)
    if year_match:
        return year_match.group()
    return None


def _check_keyword(title: str, abstract: str, keyword: Optional[str]) -> bool:
    if not keyword:
        return False
    keyword_lower = keyword.lower()
    return keyword_lower in title.lower() or keyword_lower in abstract.lower()


async def _get_citation_count_and_id(client: httpx.AsyncClient, paper_title: str) -> tuple[int, str]:
    search_url = f"https://scholar.google.com/scholar?q={paper_title.replace(' ', '+')}&hl=en&as_sdt=0,5"

    if rate_limiter.should_wait_for_captcha_cooldown():
        wait_time = rate_limiter.get_captcha_cooldown_remaining()
        if wait_time > 0:
            print(f"Waiting for captcha cooldown: {wait_time:.0f}s")
            await asyncio.sleep(wait_time)

    max_retries = 5
    for attempt in range(max_retries):
        delay = rate_limiter.get_recommended_delay()
        print(f"Request delay (attempt {attempt + 1}): {delay:.1f}s")
        await asyncio.sleep(delay)

        response = await client.get(search_url, headers=_get_headers())

        if response.status_code == 429:
            rate_limiter.record_rate_limit()
            if attempt < max_retries - 1:
                wait_time = rate_limiter.get_recommended_delay() * (attempt + 1)
                print(f"Rate limited (429), waiting {wait_time:.1f}s before retry...")
                await asyncio.sleep(wait_time)
                continue
            else:
                raise Exception("Rate limited (HTTP 429). Google Scholar is blocking this IP. Please try again later or change your IP address.")

        if _is_captcha_page(response):
            rate_limiter.record_captcha()
            if attempt < max_retries - 1:
                wait_time = rate_limiter.get_recommended_delay() * 2
                print(f"Captcha detected, waiting {wait_time:.1f}s before retry...")
                await asyncio.sleep(wait_time)
                continue
            else:
                raise Exception("Google Scholar captcha verification required. Please try again later or change your IP address.")

        rate_limiter.record_success()
        break

    soup = BeautifulSoup(response.text, "lxml")

    cited_by = 0
    citation_id = None

    for text in soup.find_all(string=re.compile(r"Cited by")):
        cited_by = _parse_citation_count(text)
        href = text.parent.get("href", "") if text.parent else ""
        id_match = re.search(r"cites=([^&]+)", href)
        if id_match:
            citation_id = id_match.group(1)
            break

    if citation_id is None:
        for result in soup.select("div.gs_r"):
            links = result.select("a")
            for link in links:
                href = link.get("href", "")
                if "cites=" in href:
                    id_match = re.search(r"cites=([^&]+)", href)
                    if id_match:
                        citation_id = id_match.group(1)
                        break
            if citation_id:
                break

    transition_delay = random.uniform(10.0, 20.0)
    print(f"Transition delay before citations request: {transition_delay:.1f}s")
    await asyncio.sleep(transition_delay)

    return cited_by, citation_id


def _parse_papers_from_page(soup: BeautifulSoup, keyword: Optional[str]) -> List[PaperInfo]:
    papers: List[PaperInfo] = []
    results = soup.select("div.gs_r")

    for result in results:
        try:
            title_elem = result.select_one("h3.gs_rt a")
            if not title_elem:
                continue

            title = title_elem.get_text(strip=True)
            url = title_elem.get("href")

            abstract_elem = result.select_one("div.gs_rs")
            abstract = abstract_elem.get_text(strip=True) if abstract_elem else ""

            info_elem = result.select_one("div.gs_a")
            meta_text = info_elem.get_text(strip=True) if info_elem else ""

            year = _extract_year(meta_text)

            parts = meta_text.split(" - ")
            publisher = None
            if parts:
                publisher = parts[0].strip() if parts[0] else None

            contains_keyword = _check_keyword(title, abstract, keyword)

            papers.append(PaperInfo(
                title=title,
                year=year,
                publisher=publisher,
                url=url,
                abstract=abstract,
                contains_keyword=contains_keyword,
            ))
        except Exception:
            continue
    return papers


async def _get_citing_papers(
    client: httpx.AsyncClient,
    citation_id: str,
    keyword: Optional[str],
    max_results: int = 50,
) -> List[PaperInfo]:
    all_papers: List[PaperInfo] = []
    start = 0
    page_num = 0
    max_retries = 3

    initial_delay = random.uniform(10.0, 20.0)
    print(f"Initial delay before fetching citations: {initial_delay:.1f}s")
    await asyncio.sleep(initial_delay)

    while len(all_papers) < max_results:
        if rate_limiter.should_wait_for_captcha_cooldown():
            remaining = rate_limiter.get_captcha_cooldown_remaining()
            print(f"Waiting for captcha cooldown: {remaining:.0f}s remaining")
            await asyncio.sleep(min(remaining, 60))

        cites_url = f"https://scholar.google.com/scholar?cites={citation_id}&as_sdt=2005&sciodt=0,5&hl=en&start={start}"

        delay = rate_limiter.get_recommended_delay()
        print(f"Request delay: {delay:.1f}s")
        await asyncio.sleep(delay)

        retry_count = 0
        response = None

        while retry_count < max_retries:
            try:
                response = await client.get(cites_url, headers=_get_headers())

                if response.status_code == 429:
                    rate_limiter.record_rate_limit()
                    wait_time = rate_limiter.get_recommended_delay()
                    print(f"Rate limited (429), waiting {wait_time:.1f}s...")
                    await asyncio.sleep(wait_time)
                    retry_count += 1
                    continue

                if _is_captcha_page(response):
                    rate_limiter.record_captcha()
                    raise Exception("Google Scholar captcha verification required. Please try again later or change your IP address.")

                if _is_rate_limited(response):
                    rate_limiter.record_rate_limit()
                    wait_time = rate_limiter.get_recommended_delay()
                    print(f"Rate limited, waiting {wait_time:.1f}s...")
                    await asyncio.sleep(wait_time)
                    retry_count += 1
                    continue

                rate_limiter.record_success()
                break

            except Exception as e:
                retry_count += 1
                if retry_count >= max_retries:
                    raise
                wait_time = rate_limiter.get_recommended_delay() * retry_count
                print(f"Error, retry {retry_count}, waiting {wait_time:.1f}s...")
                await asyncio.sleep(wait_time)

        if response is None:
            break

        soup = BeautifulSoup(response.text, "lxml")

        papers = _parse_papers_from_page(soup, keyword)
        if not papers:
            break

        all_papers.extend(papers)
        page_num += 1

        if len(papers) < 10:
            break

        start += 10
        page_delay = rate_limiter.get_adaptive_delay_for_page(page_num)
        print(f"Page {page_num} delay: {page_delay:.1f}s")
        await asyncio.sleep(page_delay)

    return all_papers[:max_results]


def _save_to_excel(papers: List[PaperInfo], output_path: str, keyword: Optional[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Citations"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    headers = ["Title", "Year", "Publisher", "Abstract", "URL", "Contains Keyword"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row, paper in enumerate(papers, 2):
        ws.cell(row=row, column=1, value=paper.title)
        ws.cell(row=row, column=2, value=paper.year or "N/A")
        ws.cell(row=row, column=3, value=paper.publisher or "N/A")
        ws.cell(row=row, column=4, value=paper.abstract if paper.abstract else "N/A")
        ws.cell(row=row, column=5, value=paper.url or "N/A")
        ws.cell(row=row, column=6, value="YES" if paper.contains_keyword else "NO")

        if paper.contains_keyword and keyword:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = highlight_fill

        ws.row_dimensions[row].height = 60

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 80
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 15

    wb.save(output_path)


@mcp.tool(
    name="scholar_search_and_export",
    annotations={
        "title": "Search Google Scholar and Export Citations to Excel",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": True,
    },
)
async def scholar_search_and_export(params: SearchInput) -> str:
    """
    Search Google Scholar for a paper and export citations to an Excel file.

    This tool searches Google Scholar using the provided paper title, retrieves
    citing papers, and exports the results to an Excel spreadsheet. Papers containing
    the specified keyword in title OR abstract will be marked.

    Args:
        params (SearchInput): Validated input parameters containing:
            - paper_title (str): Title of the paper to search
            - keyword (Optional[str]): Keyword to mark papers (checks both title AND abstract)
            - max_results (Optional[int]): Maximum number of results (default: 50, max: 100)
            - output_filename (Optional[str]): Custom output filename for the Excel file

    Returns:
        str: JSON-formatted response containing:
            - success (bool): Whether the operation succeeded
            - paper_title (str): The search query
            - cited_by_count (int): Number of papers citing the main paper
            - results_count (int): Number of citation entries found
            - papers_with_abstract (int): Number of papers that have abstract
            - keyword_matched_count (int): Number of papers containing the keyword
            - excel_file (str): Path to the saved Excel file
            - citations (list): Array of citation objects with title, year, publisher, abstract, url, contains_keyword
    """
    try:
        async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
            cited_by, citation_id = await _get_citation_count_and_id(client, params.paper_title)

            if citation_id:
                papers = await _get_citing_papers(
                    client,
                    citation_id,
                    params.keyword,
                    params.max_results,
                )
            else:
                papers = []

        keyword_matched_count = sum(1 for p in papers if p.contains_keyword)
        papers_with_abstract_count = sum(1 for p in papers if p.abstract)

        if not papers and cited_by == 0:
            return json.dumps({
                "success": True,
                "paper_title": params.paper_title,
                "cited_by_count": 0,
                "results_count": 0,
                "papers_with_abstract": 0,
                "keyword_matched_count": 0,
                "excel_file": None,
                "citations": [],
                "message": f"No citation results found for '{params.paper_title}'",
            }, indent=2)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if params.output_filename:
            safe_filename = re.sub(r"[^\w\s-]", "", params.output_filename)
            safe_filename = re.sub(r"\s+", "_", safe_filename)
            output_path = f"{safe_filename}.xlsx"
        else:
            safe_title = re.sub(r"[^\w\s-]", "", params.paper_title)[:30]
            safe_title = re.sub(r"\s+", "_", safe_title)
            output_path = f"citations_{safe_title}_{timestamp}.xlsx"

        _save_to_excel(papers, output_path, params.keyword)

        result = {
            "success": True,
            "paper_title": params.paper_title,
            "cited_by_count": cited_by,
            "results_count": len(papers),
            "papers_with_abstract": papers_with_abstract_count,
            "keyword_matched_count": keyword_matched_count,
            "keyword": params.keyword,
            "excel_file": output_path,
            "citations": [
                {
                    "title": p.title,
                    "year": p.year,
                    "publisher": p.publisher,
                    "abstract": p.abstract,
                    "url": p.url,
                    "contains_keyword": p.contains_keyword,
                }
                for p in papers
            ],
        }

        return json.dumps(result, indent=2, ensure_ascii=False)

    except httpx.HTTPStatusError as e:
        if e.response.status_code == 429:
            return json.dumps({
                "success": False,
                "error": "Rate limit exceeded (HTTP 429)",
                "message": "Google Scholar rate limit exceeded. Please wait a few minutes and try again, or change your IP address.",
            }, indent=2)
        return json.dumps({
            "success": False,
            "error": f"HTTP error occurred: {e.response.status_code}",
            "message": "Failed to access Google Scholar. Please try again later.",
        }, indent=2)
    except httpx.TimeoutException:
        return json.dumps({
            "success": False,
            "error": "Request timed out",
            "message": "Google Scholar request timed out. Please try again.",
        }, indent=2)
    except Exception as e:
        error_msg = str(e)
        if "captcha" in error_msg.lower() or "verification required" in error_msg.lower():
            return json.dumps({
                "success": False,
                "error": "Captcha verification required",
                "message": error_msg,
            }, indent=2)
        return json.dumps({
            "success": False,
            "error": str(e),
            "message": "An unexpected error occurred while searching Google Scholar.",
        }, indent=2)


@mcp.tool(
    name="scholar_get_citation_count",
    annotations={
        "title": "Get Citation Count for a Paper",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": True,
    },
)
async def scholar_get_citation_count(paper_title: str) -> str:
    """
    Get the citation count for a paper from Google Scholar.

    Args:
        paper_title (str): Title of the paper to search

    Returns:
        str: JSON-formatted response containing citation count
    """
    try:
        async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
            cited_by, _ = await _get_citation_count_and_id(client, paper_title)

        return json.dumps({
            "success": True,
            "paper_title": paper_title,
            "citation_count": cited_by,
            "message": f"'{paper_title}' has been cited by {cited_by} papers on Google Scholar.",
        }, indent=2, ensure_ascii=False)

    except httpx.HTTPStatusError as e:
        if e.response.status_code == 429:
            return json.dumps({
                "success": False,
                "error": "Rate limit exceeded (HTTP 429)",
                "message": "Google Scholar rate limit exceeded. Please wait a few minutes and try again, or change your IP address.",
            }, indent=2)
        return json.dumps({
            "success": False,
            "error": f"HTTP error occurred: {e.response.status_code}",
            "message": "Failed to access Google Scholar. Please try again later.",
        }, indent=2)
    except httpx.TimeoutException:
        return json.dumps({
            "success": False,
            "error": "Request timed out",
            "message": "Google Scholar request timed out. Please try again.",
        }, indent=2)
    except Exception as e:
        error_msg = str(e)
        if "captcha" in error_msg.lower() or "verification required" in error_msg.lower():
            return json.dumps({
                "success": False,
                "error": "Captcha verification required",
                "message": error_msg,
            }, indent=2)
        return json.dumps({
            "success": False,
            "error": str(e),
            "message": "An unexpected error occurred while searching Google Scholar.",
        }, indent=2)


if __name__ == "__main__":
    mcp.run()