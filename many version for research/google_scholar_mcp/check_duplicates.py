#!/usr/bin/env python3
from openpyxl import load_workbook

FILE1 = "/Users/hyijie/Desktop/Interest_try_test/google_scholar_mcp/我在这里我在这里_glm5.xlsx"
FILE2 = "/Users/hyijie/Desktop/Interest_try_test/google_scholar_mcp/我在这里我在这里_part2.xlsx"

def find_duplicates():
    titles1 = {}
    titles2 = {}
    
    wb1 = load_workbook(FILE1)
    ws1 = wb1.active
    for row in range(2, ws1.max_row + 1):
        title = ws1.cell(row=row, column=1).value
        if title:
            titles1[title] = row
    wb1.close()
    
    wb2 = load_workbook(FILE2)
    ws2 = wb2.active
    for row in range(2, ws2.max_row + 1):
        title = ws2.cell(row=row, column=1).value
        if title:
            titles2[title] = row
    wb2.close()
    
    duplicates = set(titles1.keys()) & set(titles2.keys())
    
    print(f"文件1 (我在这里我在这里_glm5.xlsx): {len(titles1)} 篇")
    print(f"文件2 (我在这里我在这里_part2.xlsx): {len(titles2)} 篇")
    print(f"重复论文数量: {len(duplicates)} 篇")
    print("=" * 80)
    
    if duplicates:
        print("\n去重的论文列表:")
        print("-" * 80)
        for i, title in enumerate(duplicates, 1):
            print(f"{i}. {title}")
    else:
        print("\n没有发现重复论文")

if __name__ == "__main__":
    find_duplicates()
