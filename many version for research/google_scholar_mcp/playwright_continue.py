#!/usr/bin/env python3
import asyncio
import time
import random
import re
import sys
import os

from playwright.async_api import async_playwright, Page, Browser
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

PAPER_TITLE = "A supramodal accumulation-to-bound signal that determines perceptual decisions in humans"
OUTPUT_FILENAME = "/Users/hyijie/Desktop/Interest_try_test/v1_citation_from_google_part2.xlsx"
CITATION_ID = "11385087487892509364"
START_PAGE = 42
MAX_RESULTS = 500


async def human_like_delay():
    delays = [
        (1.0, 2.0),
        (2.0, 4.0),
        (3.0, 6.0),
        (1.5, 3.5),
    ]
    min_d, max_d = random.choice(delays)
    await asyncio.sleep(random.uniform(min_d, max_d))


def extract_year(text: str) -> str | None:
    year_match = re.search(r"\b(19|20)\d{2}\b", text)
    if year_match:
        return year_match.group()
    return None


async def is_captcha_page(page: Page) -> bool:
    try:
        current_url = page.url.lower()
        
        if "sorry" in current_url:
            return True
        if "verify" in current_url:
            return True
        
        try:
            content = await page.content()
            content_lower = content.lower()
            
            if "sorry" in content_lower and ("captcha" in content_lower or "unusual traffic" in content_lower):
                return True
            if "recaptcha" in content_lower:
                return True
        except Exception:
            pass
        
        return False
    except Exception:
        return False


async def wait_for_captcha_solve(page: Page, timeout: int = 300):
    print("\n" + "=" * 60)
    print("CAPTCHA DETECTED!")
    print("Please solve the captcha in the browser window.")
    print("The script will continue automatically after you solve it.")
    print("=" * 60 + "\n")
    
    start_time = time.time()
    while time.time() - start_time < timeout:
        try:
            if not await is_captcha_page(page):
                print("Captcha solved! Continuing...")
                await asyncio.sleep(2)
                return True
        except Exception:
            pass
        await asyncio.sleep(2)
    
    return False


async def get_citing_papers(page: Page, citation_id: str, start_page: int, max_results: int = 100) -> list[dict]:
    all_papers = []
    start = start_page * 10
    page_num = start_page
    
    while len(all_papers) < max_results:
        cites_url = f"https://scholar.google.com/scholar?cites={citation_id}&as_sdt=2005&sciodt=0,5&hl=en&start={start}"
        
        print(f"\nFetching page {page_num + 1}...")
        
        try:
            await page.goto(cites_url, wait_until="networkidle", timeout=60000)
        except Exception as e:
            print(f"Navigation error: {e}")
            await asyncio.sleep(5)
            continue
        
        await human_like_delay()
        
        if await is_captcha_page(page):
            if not await wait_for_captcha_solve(page):
                print("Could not solve captcha, stopping...")
                break
        
        results = await page.locator("div.gs_r").all()
        
        if not results:
            print("No more results found")
            break
        
        for result in results:
            try:
                title_elem = result.locator("h3.gs_rt a")
                if not await title_elem.count():
                    continue
                    
                title = await title_elem.text_content() or ""
                title = title.strip()
                url = await title_elem.get_attribute("href") or ""
                
                abstract = ""
                abstract_elem = result.locator("div.gs_rs")
                if await abstract_elem.count():
                    abstract = await abstract_elem.text_content() or ""
                    abstract = abstract.strip()
                
                meta_text = ""
                info_elem = result.locator("div.gs_a")
                if await info_elem.count():
                    meta_text = await info_elem.text_content() or ""
                    meta_text = meta_text.strip()
                
                year = extract_year(meta_text)
                
                publisher = None
                if meta_text:
                    parts = meta_text.split(" - ")
                    if parts:
                        publisher = parts[0].strip()
                
                all_papers.append({
                    "title": title,
                    "year": year,
                    "publisher": publisher,
                    "abstract": abstract,
                    "url": url,
                })
                
            except Exception:
                continue
        
        print(f"Found {len(results)} papers on page {page_num + 1}, total: {len(all_papers)}")
        
        if len(results) < 10:
            break
        
        page_num += 1
        start += 10
        
        if page_num % 2 == 0:
            batch_delay = random.uniform(8, 15)
            print(f"Batch rest: {batch_delay:.1f}s...")
            await asyncio.sleep(batch_delay)
        else:
            await asyncio.sleep(random.uniform(3, 6))
    
    return all_papers[:max_results]


def save_to_excel(papers: list[dict], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Citations"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    headers = ["Title", "Year", "Publisher", "Abstract", "URL"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    for row, paper in enumerate(papers, 2):
        ws.cell(row=row, column=1, value=paper["title"])
        ws.cell(row=row, column=2, value=paper["year"] or "N/A")
        ws.cell(row=row, column=3, value=paper["publisher"] or "N/A")
        ws.cell(row=row, column=4, value=paper["abstract"] if paper["abstract"] else "N/A")
        ws.cell(row=row, column=5, value=paper["url"] or "N/A")
        ws.row_dimensions[row].height = 60
    
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 80
    ws.column_dimensions["E"].width = 40
    
    wb.save(output_path)
    print(f"\nSaved {len(papers)} papers to {output_path}")


async def main():
    print("=" * 60)
    print(f"Continuing from page {START_PAGE + 1}")
    print(f"Output file: {OUTPUT_FILENAME}")
    print("=" * 60)
    
    async with async_playwright() as p:
        print("\nInitializing browser...")
        browser = await p.chromium.launch(
            headless=False,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
                "--disable-dev-shm-usage",
            ]
        )
        
        context = await browser.new_context(
            viewport={"width": 1920, "height": 1080},
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            locale="en-US",
        )
        
        page = await context.new_page()
        
        try:
            print("\nFetching remaining citing papers...")
            papers = await get_citing_papers(page, CITATION_ID, START_PAGE, MAX_RESULTS)
            
            print(f"\nTotal papers collected: {len(papers)}")
            
            output_path = f"/Users/hyijie/Desktop/Interest_try_test/google_scholar_mcp/{OUTPUT_FILENAME}"
            save_to_excel(papers, output_path)
            
            print("\nDone!")
            
        except Exception as e:
            print(f"Error: {e}")
            import traceback
            traceback.print_exc()
        finally:
            print("\nClosing browser...")
            await browser.close()


if __name__ == "__main__":
    asyncio.run(main())
