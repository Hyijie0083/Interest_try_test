#!/usr/bin/env python3
"""
Playwright script to search Google Scholar and extract detailed information.
This script opens each paper's page to get FULL abstract and keywords.
Saves progress after each paper.
"""

import asyncio
import time
import random
import re
import sys
import os

from playwright.async_api import async_playwright, Page, Browser
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

PAPER_TITLE = "A supramodal accumulation-to-bound signal that determines perceptual decisions in humans"
OUTPUT_FILENAME = "/Users/hyijie/Desktop/Interest_try_test/v1_citation_from_google_part3.xlsx"
MAX_RESULTS = 100
START_PAGE = 86
CITATION_ID = "11385087487892509364"


def save_to_excel(papers: list[dict], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Citations"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    headers = ["Title", "Year", "Publisher", "Abstract", "Keywords", "URL"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    for row, paper in enumerate(papers, 2):
        ws.cell(row=row, column=1, value=paper["title"])
        ws.cell(row=row, column=2, value=paper["year"] or "N/A")
        ws.cell(row=row, column=3, value=paper["publisher"] or "N/A")
        ws.cell(row=row, column=4, value=paper["abstract"] if paper["abstract"] else "N/A")
        ws.cell(row=row, column=5, value=paper["keywords"] if paper["keywords"] else "N/A")
        ws.cell(row=row, column=6, value=paper["url"] or "N/A")
        ws.row_dimensions[row].height = 100
    
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 80
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 40
    
    wb.save(output_path)


async def human_like_delay():
    await asyncio.sleep(random.uniform(1.5, 3.5))


def extract_year(text: str) -> str | None:
    year_match = re.search(r"\b(19|20)\d{2}\b", text)
    if year_match:
        return year_match.group()
    return None


def parse_citation_count(text: str) -> int:
    match = re.search(r"[\d,]+", text.replace(",", ""))
    if match:
        return int(match.group().replace(",", ""))
    return 0


async def is_captcha_page(page: Page) -> bool:
    try:
        current_url = page.url.lower()
        if "sorry" in current_url or "verify" in current_url:
            return True
        try:
            content = await page.content()
            content_lower = content.lower()
            if "sorry" in content_lower and ("captcha" in content_lower or "unusual traffic" in content_lower):
                return True
            if "recaptcha" in content_lower:
                return True
        except Exception:
            pass
        return False
    except Exception:
        return False


async def is_journal_verification_page(page: Page) -> bool:
    try:
        content = await page.content()
        content_lower = content.lower()
        
        captcha_iframe_patterns = [
            'iframe.*recaptcha',
            'iframe.*hcaptcha',
            'iframe.*captcha',
            'g-recaptcha',
            'h-captcha',
            'cf-turnstile',
            'challenge-platform',
            'cf-browser-verification',
        ]
        
        for pattern in captcha_iframe_patterns:
            if re.search(pattern, content_lower):
                print(f"    Detected captcha/challenge element: {pattern}")
                return True
        
        cloudflare_indicators = [
            'checking your browser',
            'please wait while we verify',
            'just a moment',
            'enable javascript and cookies',
        ]
        
        for indicator in cloudflare_indicators:
            if indicator in content_lower:
                body_text = ""
                try:
                    body = page.locator('body')
                    body_text = await body.text_content() or ""
                    body_text = body_text.lower()
                except:
                    pass
                
                if len(body_text) < 500:
                    print(f"    Detected Cloudflare challenge: '{indicator}'")
                    return True
        
        hard_block_patterns = [
            '403 forbidden',
            'access denied',
            'your access has been blocked',
        ]
        
        for pattern in hard_block_patterns:
            if pattern in content_lower:
                try:
                    title = await page.title()
                    if title and any(p in title.lower() for p in ['error', 'denied', 'forbidden', '403']):
                        print(f"    Detected hard block: '{pattern}'")
                        return True
                except:
                    pass
        
        return False
    except Exception:
        return False


async def wait_for_captcha_solve(page: Page, timeout: int = 300):
    print("\n" + "=" * 60)
    print("CAPTCHA DETECTED!")
    print("Please solve the captcha in the browser window.")
    print("=" * 60 + "\n")
    
    start_time = time.time()
    while time.time() - start_time < timeout:
        try:
            if not await is_captcha_page(page):
                print("Captcha solved! Continuing...")
                await asyncio.sleep(2)
                return True
        except Exception:
            pass
        await asyncio.sleep(2)
    return False


async def get_paper_detail(page: Page, paper_url: str) -> tuple[str, str, bool]:
    if not paper_url or not paper_url.startswith("http"):
        return "", "", False
    
    abstract = ""
    keywords = ""
    skipped = False
    
    try:
        new_page = await page.context.new_page()
        try:
            print(f"    Opening paper page: {paper_url[:60]}...")
            await new_page.goto(paper_url, wait_until="domcontentloaded", timeout=45000)
            await asyncio.sleep(2)
            
            if await is_journal_verification_page(new_page):
                print(f"    VERIFICATION REQUIRED - Skipping this journal")
                skipped = True
                return "", "", True
            
            try:
                meta_keywords = await new_page.locator('meta[name="keywords"]').get_attribute("content")
                if meta_keywords and len(meta_keywords) > 3:
                    keywords = meta_keywords.strip()
                    print(f"    Found keywords: {keywords[:50]}...")
            except Exception:
                pass
            
            try:
                meta_desc = await new_page.locator('meta[name="description"]').get_attribute("content")
                if meta_desc and len(meta_desc) > 200:
                    abstract = meta_desc.strip()
                    print(f"    Found abstract from meta description: {len(abstract)} chars")
            except Exception:
                pass
            
            if not abstract or len(abstract) < 200:
                abstract_selectors = [
                    "abstract",
                    ".abstract",
                    "#abstract",
                    ".article-abstract",
                    ".abstract-text",
                    "[class*='abstract']",
                    "section.abstract",
                    "div.abstract",
                ]
                for selector in abstract_selectors:
                    try:
                        elem = new_page.locator(selector).first
                        if await elem.count():
                            text = await elem.text_content()
                            if text and len(text.strip()) > 200:
                                abstract = text.strip()
                                print(f"    Found abstract from '{selector}': {len(abstract)} chars")
                                break
                    except Exception:
                        continue
            
            if not abstract or len(abstract) < 200:
                try:
                    paragraphs = await new_page.locator("p, div.paragraph, section p").all()
                    for p in paragraphs[:30]:
                        text = await p.text_content()
                        if text and len(text) > 300:
                            text_lower = text.lower()
                            if "abstract" in text_lower[:150] or "we " in text_lower[:50] or "this " in text_lower[:50]:
                                abstract = text.strip()
                                print(f"    Found abstract from paragraph: {len(abstract)} chars")
                                break
                except Exception:
                    pass
            
            if not abstract:
                print(f"    Could not find full abstract on page")
            
        finally:
            await new_page.close()
    except Exception as e:
        print(f"    Error getting detail: {e}")
    
    return abstract[:5000] if abstract else "", keywords, skipped


async def search_paper(page: Page, paper_title: str) -> tuple[int, str | None]:
    search_url = f"https://scholar.google.com/scholar?q={paper_title.replace(' ', '+')}&hl=en&as_sdt=0,5"
    
    print(f"Navigating to Google Scholar...")
    await page.goto(search_url, wait_until="networkidle", timeout=60000)
    await human_like_delay()
    
    if await is_captcha_page(page):
        if not await wait_for_captcha_solve(page):
            return 0, None
    
    cited_by = 0
    citation_id = None
    
    try:
        cited_by_links = await page.get_by_text("Cited by").all()
        if cited_by_links:
            link = cited_by_links[0]
            href = await link.get_attribute("href")
            if href and "cites=" in href:
                id_match = re.search(r"cites=([^&]+)", href)
                if id_match:
                    citation_id = id_match.group(1)
            
            cited_text = await link.text_content()
            cited_by = parse_citation_count(cited_text or "")
            print(f"Found: Cited by {cited_by} papers")
    except Exception as e:
        print(f"No 'Cited by' link found: {e}")
    
    return cited_by, citation_id


async def get_citing_papers_with_details(page: Page, citation_id: str, max_results: int, output_path: str, start_page: int = 0) -> list[dict]:
    all_papers = []
    start = start_page * 10
    page_num = start_page
    
    while len(all_papers) < max_results:
        cites_url = f"https://scholar.google.com/scholar?cites={citation_id}&as_sdt=2005&sciodt=0,5&hl=en&start={start}"
        
        print(f"\nFetching page {page_num + 1}...")
        
        try:
            await page.goto(cites_url, wait_until="networkidle", timeout=60000)
        except Exception as e:
            print(f"Navigation error: {e}")
            await asyncio.sleep(5)
            continue
        
        await human_like_delay()
        
        if await is_captcha_page(page):
            if not await wait_for_captcha_solve(page):
                print("Could not solve captcha, stopping...")
                break
        
        results = await page.locator("div.gs_r").all()
        
        if not results:
            print("No more results found")
            break
        
        for i, result in enumerate(results):
            if len(all_papers) >= max_results:
                break
                
            try:
                title_elem = result.locator("h3.gs_rt a")
                if not await title_elem.count():
                    continue
                    
                title = await title_elem.text_content() or ""
                title = title.strip()
                url = await title_elem.get_attribute("href") or ""
                
                meta_text = ""
                info_elem = result.locator("div.gs_a")
                if await info_elem.count():
                    meta_text = await info_elem.text_content() or ""
                    meta_text = meta_text.strip()
                
                year = extract_year(meta_text)
                
                publisher = None
                if meta_text:
                    parts = meta_text.split(" - ")
                    if parts:
                        publisher = parts[0].strip()
                
                print(f"\n  Paper {len(all_papers) + 1}/{max_results}: {title[:50]}...")
                
                detail_abstract, keywords, skipped = await get_paper_detail(page, url)
                
                if skipped:
                    print(f"  SKIPPED due to verification - Saving basic info only")
                    paper = {
                        "title": title,
                        "year": year,
                        "publisher": publisher,
                        "abstract": "",
                        "keywords": "",
                        "url": url,
                    }
                else:
                    paper = {
                        "title": title,
                        "year": year,
                        "publisher": publisher,
                        "abstract": detail_abstract,
                        "keywords": keywords,
                        "url": url,
                    }
                
                all_papers.append(paper)
                
                print(f"  SAVING to {output_path}...")
                save_to_excel(all_papers, output_path)
                if skipped:
                    print(f"  SAVED! (Skipped due to verification)")
                else:
                    print(f"  SAVED! (Abstract: {len(detail_abstract)} chars, Keywords: {'Yes' if keywords else 'No'})")
                
                await asyncio.sleep(random.uniform(2, 4))
                
            except Exception as e:
                print(f"Error processing paper: {e}")
                continue
        
        print(f"\nPage {page_num + 1} complete. Total: {len(all_papers)} papers")
        
        if len(results) < 10:
            break
        
        page_num += 1
        start += 10
        
        await asyncio.sleep(random.uniform(3, 6))
    
    return all_papers


async def main():
    print("=" * 60)
    print(f"Searching for: {PAPER_TITLE}")
    print(f"Output file: {OUTPUT_FILENAME}")
    print(f"Starting from page: {START_PAGE + 1}")
    print("This script will open each paper's page to get FULL abstract.")
    print("Progress is saved after each paper!")
    print("=" * 60)
    
    async with async_playwright() as p:
        print("\nInitializing browser...")
        browser = await p.chromium.launch(
            headless=False,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
                "--disable-dev-shm-usage",
            ]
        )
        
        context = await browser.new_context(
            viewport={"width": 1920, "height": 1080},
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            locale="en-US",
        )
        
        page = await context.new_page()
        
        try:
            citation_id = CITATION_ID
            
            print(f"\nCitation ID: {citation_id}")
            print(f"Starting from page {START_PAGE + 1}...")
            
            print("\nFetching citing papers with FULL details...")
            papers = await get_citing_papers_with_details(page, citation_id, MAX_RESULTS, OUTPUT_FILENAME, START_PAGE)
            
            print(f"\nTotal papers collected: {len(papers)}")
            print(f"\nDone! Saved to {OUTPUT_FILENAME}")
            
        except Exception as e:
            print(f"Error: {e}")
            import traceback
            traceback.print_exc()
        finally:
            print("\nClosing browser...")
            await browser.close()


if __name__ == "__main__":
    asyncio.run(main())
