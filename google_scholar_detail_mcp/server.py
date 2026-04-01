#!/usr/bin/env python3
"""
MCP Server for Google Scholar with Detail Extraction.

This server provides tools to search Google Scholar, retrieve citations,
and export citation data to Excel files with detailed abstract and keywords.
"""

import asyncio
import json
import random
import re
import time
from datetime import datetime
from typing import List, Optional

import httpx
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, ConfigDict, Field, field_validator

from mcp.server.fastmcp import FastMCP

mcp = FastMCP("google_scholar_detail_mcp")

USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.2 Safari/605.1.15",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0",
]

HEADERS_TEMPLATE = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Sec-Fetch-User": "?1",
    "Cache-Control": "max-age=0",
    "TE": "trailers",
}


def _get_headers() -> dict:
    headers = HEADERS_TEMPLATE.copy()
    headers["User-Agent"] = random.choice(USER_AGENTS)
    return headers


def _parse_citation_count(text: str) -> int:
    match = re.search(r"[\d,]+", text.replace(",", ""))
    if match:
        return int(match.group().replace(",", ""))
    return 0


def _extract_year(text: str) -> Optional[str]:
    year_match = re.search(r"\b(19|20)\d{2}\b", text)
    if year_match:
        return year_match.group()
    return None


def _check_keyword(title: str, abstract: str, keyword: Optional[str]) -> bool:
    if not keyword:
        return False
    keyword_lower = keyword.lower()
    return keyword_lower in title.lower() or keyword_lower in abstract.lower()


class RateLimiter:
    def __init__(self):
        self.request_times: List[float] = []
        self.consecutive_errors: int = 0
        self.last_captcha_time: float = 0
        self.base_delay: float = 5.0
        self.max_delay: float = 120.0

    def record_success(self):
        self.consecutive_errors = 0
        self.request_times.append(time.time())
        if len(self.request_times) > 20:
            self.request_times.pop(0)

    def record_error(self):
        self.consecutive_errors += 1

    def record_captcha(self):
        self.consecutive_errors = 15
        self.last_captcha_time = time.time()

    def get_recommended_delay(self) -> float:
        if self.consecutive_errors == 0:
            return random.uniform(self.base_delay, self.base_delay * 1.5)
        delay = self.base_delay * (2 ** min(self.consecutive_errors, 6))
        return min(delay + random.uniform(0, delay * 0.5), self.max_delay)


rate_limiter = RateLimiter()


def _is_captcha_page(response: httpx.Response) -> bool:
    if "sorry" in str(response.url).lower():
        return True
    if response.status_code == 429:
        return True
    text_lower = response.text.lower()
    if "sorry" in text_lower and ("captcha" in text_lower or "unusual traffic" in text_lower):
        return True
    return False


class PaperInfo(BaseModel):
    title: str
    year: Optional[str] = None
    publisher: Optional[str] = None
    url: Optional[str] = None
    abstract: Optional[str] = None
    keywords: Optional[str] = None
    contains_keyword: bool = False


class SearchInput(BaseModel):
    model_config = ConfigDict(
        str_strip_whitespace=True,
        validate_assignment=True,
    )

    paper_title: str = Field(
        ...,
        description="Title of the paper to search on Google Scholar",
        min_length=1,
        max_length=500,
    )
    keyword: Optional[str] = Field(
        default=None,
        description="Keyword to mark papers",
        max_length=100,
    )
    max_results: Optional[int] = Field(
        default=50,
        description="Maximum number of citation results to retrieve",
        ge=1,
        le=100,
    )
    output_filename: Optional[str] = Field(
        default=None,
        description="Custom output filename for the Excel file",
        max_length=200,
    )

    @field_validator("paper_title")
    @classmethod
    def validate_paper_title(cls, v: str) -> str:
        if not v.strip():
            raise ValueError("Paper title cannot be empty")
        return v.strip()


def _save_to_excel(papers: List[PaperInfo], output_path: str, keyword: Optional[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Citations"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    headers = ["Title", "Year", "Publisher", "Abstract", "Keywords", "URL", "Contains Keyword"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row, paper in enumerate(papers, 2):
        ws.cell(row=row, column=1, value=paper.title)
        ws.cell(row=row, column=2, value=paper.year or "N/A")
        ws.cell(row=row, column=3, value=paper.publisher or "N/A")
        ws.cell(row=row, column=4, value=paper.abstract if paper.abstract else "N/A")
        ws.cell(row=row, column=5, value=paper.keywords if paper.keywords else "N/A")
        ws.cell(row=row, column=6, value=paper.url or "N/A")
        ws.cell(row=row, column=7, value="YES" if paper.contains_keyword else "NO")

        if paper.contains_keyword and keyword:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = highlight_fill

        ws.row_dimensions[row].height = 60

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 15

    wb.save(output_path)


async def _get_citation_count_and_id(client: httpx.AsyncClient, paper_title: str) -> tuple[int, str]:
    search_url = f"https://scholar.google.com/scholar?q={paper_title.replace(' ', '+')}&hl=en&as_sdt=0,5"

    delay = rate_limiter.get_recommended_delay()
    await asyncio.sleep(delay)

    response = await client.get(search_url, headers=_get_headers())

    if _is_captcha_page(response):
        rate_limiter.record_captcha()
        raise Exception("Google Scholar captcha verification required.")

    rate_limiter.record_success()

    soup = BeautifulSoup(response.text, "lxml")

    cited_by = 0
    citation_id = None

    for text in soup.find_all(string=re.compile(r"Cited by")):
        cited_by = _parse_citation_count(text)
        href = text.parent.get("href", "") if text.parent else ""
        id_match = re.search(r"cites=([^&]+)", href)
        if id_match:
            citation_id = id_match.group(1)
            break

    return cited_by, citation_id or ""


def _parse_papers_from_page(soup: BeautifulSoup, keyword: Optional[str]) -> List[PaperInfo]:
    papers: List[PaperInfo] = []
    results = soup.select("div.gs_r")

    for result in results:
        try:
            title_elem = result.select_one("h3.gs_rt a")
            if not title_elem:
                continue

            title = title_elem.get_text(strip=True)
            url = title_elem.get("href")

            abstract_elem = result.select_one("div.gs_rs")
            abstract = abstract_elem.get_text(strip=True) if abstract_elem else ""

            info_elem = result.select_one("div.gs_a")
            meta_text = info_elem.get_text(strip=True) if info_elem else ""

            year = _extract_year(meta_text)

            parts = meta_text.split(" - ")
            publisher = parts[0].strip() if parts else None

            contains_keyword = _check_keyword(title, abstract, keyword)

            papers.append(PaperInfo(
                title=title,
                year=year,
                publisher=publisher,
                url=url,
                abstract=abstract,
                keywords="",
                contains_keyword=contains_keyword,
            ))
        except Exception:
            continue
    return papers


async def _get_paper_detail(client: httpx.AsyncClient, paper_url: str) -> tuple[str, str]:
    if not paper_url or not paper_url.startswith("http"):
        return "", ""
    
    delay = rate_limiter.get_recommended_delay()
    await asyncio.sleep(delay)
    
    try:
        response = await client.get(paper_url, headers=_get_headers(), timeout=30.0)
        
        if response.status_code != 200:
            return "", ""
        
        soup = BeautifulSoup(response.text, "lxml")
        
        abstract = ""
        keywords = ""
        
        meta_desc = soup.find("meta", attrs={"name": "description"})
        if meta_desc:
            abstract = meta_desc.get("content", "")
        
        meta_keywords = soup.find("meta", attrs={"name": "keywords"})
        if meta_keywords:
            keywords = meta_keywords.get("content", "")
        
        rate_limiter.record_success()
        return abstract[:2000] if abstract else "", keywords
        
    except Exception:
        rate_limiter.record_error()
        return "", ""


async def _get_citing_papers(
    client: httpx.AsyncClient,
    citation_id: str,
    keyword: Optional[str],
    max_results: int = 50,
) -> List[PaperInfo]:
    all_papers: List[PaperInfo] = []
    start = 0
    max_retries = 3

    while len(all_papers) < max_results:
        cites_url = f"https://scholar.google.com/scholar?cites={citation_id}&as_sdt=2005&sciodt=0,5&hl=en&start={start}"

        delay = rate_limiter.get_recommended_delay()
        await asyncio.sleep(delay)

        retry_count = 0
        response = None

        while retry_count < max_retries:
            try:
                response = await client.get(cites_url, headers=_get_headers())

                if _is_captcha_page(response):
                    rate_limiter.record_captcha()
                    raise Exception("Captcha verification required")

                rate_limiter.record_success()
                break

            except Exception:
                retry_count += 1
                if retry_count >= max_retries:
                    return all_papers
                await asyncio.sleep(delay * retry_count)

        if response is None:
            break

        soup = BeautifulSoup(response.text, "lxml")
        papers = _parse_papers_from_page(soup, keyword)
        
        if not papers:
            break

        all_papers.extend(papers)
        
        if len(papers) < 10:
            break

        start += 10

    return all_papers[:max_results]


@mcp.tool(
    name="scholar_search_and_export",
    annotations={
        "title": "Search Google Scholar and Export Citations to Excel",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": True,
    },
)
async def scholar_search_and_export(params: SearchInput) -> str:
    """
    Search Google Scholar for a paper and export citations to an Excel file.
    """
    try:
        async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
            cited_by, citation_id = await _get_citation_count_and_id(client, params.paper_title)

            if citation_id:
                papers = await _get_citing_papers(
                    client,
                    citation_id,
                    params.keyword,
                    params.max_results,
                )
            else:
                papers = []

        keyword_matched_count = sum(1 for p in papers if p.contains_keyword)

        if not papers and cited_by == 0:
            return json.dumps({
                "success": True,
                "paper_title": params.paper_title,
                "cited_by_count": 0,
                "results_count": 0,
                "keyword_matched_count": 0,
                "excel_file": None,
                "message": f"No citation results found for '{params.paper_title}'",
            }, indent=2)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if params.output_filename:
            safe_filename = re.sub(r"[^\w\s-]", "", params.output_filename)
            safe_filename = re.sub(r"\s+", "_", safe_filename)
            output_path = f"{safe_filename}.xlsx"
        else:
            safe_title = re.sub(r"[^\w\s-]", "", params.paper_title)[:30]
            safe_title = re.sub(r"\s+", "_", safe_title)
            output_path = f"citations_{safe_title}_{timestamp}.xlsx"

        _save_to_excel(papers, output_path, params.keyword)

        result = {
            "success": True,
            "paper_title": params.paper_title,
            "cited_by_count": cited_by,
            "results_count": len(papers),
            "keyword_matched_count": keyword_matched_count,
            "keyword": params.keyword,
            "excel_file": output_path,
            "citations": [
                {
                    "title": p.title,
                    "year": p.year,
                    "publisher": p.publisher,
                    "abstract": p.abstract,
                    "keywords": p.keywords,
                    "url": p.url,
                    "contains_keyword": p.contains_keyword,
                }
                for p in papers
            ],
        }

        return json.dumps(result, indent=2, ensure_ascii=False)

    except Exception as e:
        return json.dumps({
            "success": False,
            "error": str(e),
            "message": "An error occurred while searching Google Scholar.",
        }, indent=2)


if __name__ == "__main__":
    mcp.run()
